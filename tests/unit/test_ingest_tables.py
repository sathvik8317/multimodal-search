import json

import openpyxl
import pytest

from mmsearch import config
from mmsearch.clients.fakes import FakeEmbeddingClient
from mmsearch.ingest.tables import ingest_table
from mmsearch.schema import Modality, TextSource
from tests.fixtures import CORPUS_DIR


def test_ingest_table_returns_row_with_correct_id():
    csv_path = CORPUS_DIR / "data" / "latency.csv"
    row = ingest_table(csv_path, CORPUS_DIR, FakeEmbeddingClient())
    assert row.id == "tbl:data/latency.csv"


def test_ingest_table_sets_modality_and_text_source():
    csv_path = CORPUS_DIR / "data" / "latency.csv"
    row = ingest_table(csv_path, CORPUS_DIR, FakeEmbeddingClient())
    assert row.modality == Modality.TABLE
    assert row.text_source == TextSource.TABLE_MARKDOWN


def test_ingest_table_content_text_contains_headers_and_looks_like_markdown():
    csv_path = CORPUS_DIR / "data" / "latency.csv"
    row = ingest_table(csv_path, CORPUS_DIR, FakeEmbeddingClient())
    assert "endpoint" in row.content_text
    assert "p50_ms" in row.content_text
    assert "p95_ms" in row.content_text
    assert "p99_ms" in row.content_text
    lines = row.content_text.strip().splitlines()
    assert lines[0].startswith("|")
    assert set(lines[1].replace("|", "").strip()) <= {"-", " "}


def test_ingest_table_vector_has_correct_dimension():
    csv_path = CORPUS_DIR / "data" / "latency.csv"
    row = ingest_table(csv_path, CORPUS_DIR, FakeEmbeddingClient())
    assert row.vector_cohere is None
    assert len(row.vector_openai) == config.OPENAI_EMBED_DIM


def test_ingest_table_thumbnail_ref_is_empty():
    csv_path = CORPUS_DIR / "data" / "latency.csv"
    row = ingest_table(csv_path, CORPUS_DIR, FakeEmbeddingClient())
    assert row.thumbnail_ref == ""


def test_ingest_table_source_path_is_corpus_relative():
    csv_path = CORPUS_DIR / "data" / "latency.csv"
    row = ingest_table(csv_path, CORPUS_DIR, FakeEmbeddingClient())
    assert row.source_path == "data/latency.csv"


def test_ingest_table_metadata_has_correct_row_and_column_counts():
    csv_path = CORPUS_DIR / "data" / "latency.csv"
    row = ingest_table(csv_path, CORPUS_DIR, FakeEmbeddingClient())
    metadata = json.loads(row.metadata)
    assert metadata["n_rows"] == 3
    assert metadata["n_cols"] == 4
    assert metadata["columns"] == ["endpoint", "p50_ms", "p95_ms", "p99_ms"]


# --- row cap (MAX_TABLE_ROWS) -------------------------------------------------------------

def _write_csv(path, n_data_rows: int) -> None:
    lines = ["id,value"]
    lines += [f"{i},{i * 2}" for i in range(n_data_rows)]
    path.write_text("\n".join(lines) + "\n")


def test_ingest_table_under_cap_is_unaffected(tmp_path):
    csv_path = tmp_path / "small.csv"
    _write_csv(csv_path, 5)

    row = ingest_table(csv_path, tmp_path, FakeEmbeddingClient())

    metadata = json.loads(row.metadata)
    assert metadata["n_rows"] == 5
    assert metadata["total_rows"] == 5
    assert metadata["truncated"] is False

    data_lines = row.content_text.strip().splitlines()[2:]  # skip header + separator
    assert len(data_lines) == 5


def test_ingest_table_over_cap_is_truncated(tmp_path):
    csv_path = tmp_path / "big.csv"
    _write_csv(csv_path, 250)  # over config.MAX_TABLE_ROWS (200)

    row = ingest_table(csv_path, tmp_path, FakeEmbeddingClient())

    metadata = json.loads(row.metadata)
    assert metadata["truncated"] is True
    assert metadata["total_rows"] == 250
    assert metadata["n_rows"] == config.MAX_TABLE_ROWS

    data_lines = row.content_text.strip().splitlines()[2:]  # skip header + separator
    assert len(data_lines) == config.MAX_TABLE_ROWS


def test_ingest_table_truncation_keeps_the_first_rows(tmp_path):
    csv_path = tmp_path / "big.csv"
    _write_csv(csv_path, 250)

    row = ingest_table(csv_path, tmp_path, FakeEmbeddingClient())

    assert "| 0 | 0 |" in row.content_text  # first data row present
    assert "| 199 | 398 |" in row.content_text  # 200th data row (index 199) present
    assert "| 200 | 400 |" not in row.content_text  # 201st data row absent


# --- char budget (MAX_TABLE_EMBED_CHARS) ---------------------------------------------------
#
# A wide table (many columns) can blow OpenAI's 8192-token input limit well
# before hitting MAX_TABLE_ROWS -- this is the real bug the migration
# introduced (see EMBEDDING_MIGRATION_PLAN.md): a 200-row cap sized for
# Cohere's much larger context silently produced embed inputs OpenAI
# rejected for 3 of 4 real corpus CSVs. These tests build a table wide
# enough that the char budget, not the row cap, is what binds.

def _write_wide_csv(path, n_data_rows: int, n_cols: int = 20, cell_width: int = 15) -> None:
    columns = [f"col{i}" for i in range(n_cols)]
    lines = [",".join(columns)]
    for row_i in range(n_data_rows):
        cell = "x" * cell_width
        lines.append(",".join(f"{cell}{row_i}" for _ in range(n_cols)))
    path.write_text("\n".join(lines) + "\n")


def test_ingest_table_wide_csv_is_truncated_by_char_budget_before_row_cap(tmp_path):
    csv_path = tmp_path / "wide.csv"
    _write_wide_csv(csv_path, n_data_rows=150, n_cols=20, cell_width=15)  # well under MAX_TABLE_ROWS

    row = ingest_table(csv_path, tmp_path, FakeEmbeddingClient())

    metadata = json.loads(row.metadata)
    assert metadata["truncated"] is True
    assert metadata["n_rows"] < 150  # char budget bound before the row cap did
    assert len(row.content_text) <= config.MAX_TABLE_EMBED_CHARS


def test_ingest_table_char_budget_always_keeps_at_least_one_row(tmp_path):
    csv_path = tmp_path / "huge_row.csv"
    # A single row alone exceeds MAX_TABLE_EMBED_CHARS -- must still be kept,
    # not dropped to zero rows (which would also violate content_text
    # must-not-be-empty).
    _write_wide_csv(csv_path, n_data_rows=1, n_cols=20, cell_width=1000)

    row = ingest_table(csv_path, tmp_path, FakeEmbeddingClient())

    metadata = json.loads(row.metadata)
    assert metadata["n_rows"] == 1
    assert metadata["truncated"] is False  # the only row available was kept, nothing was dropped
    data_lines = row.content_text.strip().splitlines()[2:]
    assert len(data_lines) == 1


def test_ingest_table_narrow_csv_under_row_cap_is_unaffected_by_char_budget(tmp_path):
    # Sanity check: the existing narrow-CSV row-cap tests above must still
    # bind on MAX_TABLE_ROWS, not on the new char budget.
    csv_path = tmp_path / "narrow_big.csv"
    _write_csv(csv_path, 250)

    row = ingest_table(csv_path, tmp_path, FakeEmbeddingClient())

    metadata = json.loads(row.metadata)
    assert metadata["n_rows"] == config.MAX_TABLE_ROWS
    assert len(row.content_text) <= config.MAX_TABLE_EMBED_CHARS


def test_ingest_table_is_deterministic():
    csv_path = CORPUS_DIR / "data" / "latency.csv"
    row1 = ingest_table(csv_path, CORPUS_DIR, FakeEmbeddingClient())
    row2 = ingest_table(csv_path, CORPUS_DIR, FakeEmbeddingClient())
    assert row1.id == row2.id
    assert row1.content_text == row2.content_text
    assert row1.vector_openai == row2.vector_openai


# --- .xlsx support -----------------------------------------------------------------------
#
# xlsx feeds the exact same _rows_to_markdown/_select_embedded_rows path CSV
# does (see ingest/tables.py) -- these tests establish parity against the
# equivalent CSV content, plus xlsx-specific cell-typing behavior
# (openpyxl yields typed/None values; CSV's csv.reader always yields str).

def _write_xlsx(path, header: list[str], data_rows: list[list]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(header)
    for row in data_rows:
        ws.append(row)
    wb.save(path)


def test_ingest_table_xlsx_returns_row_with_correct_id(tmp_path):
    xlsx_path = tmp_path / "latency.xlsx"
    _write_xlsx(xlsx_path, ["endpoint", "p50_ms"], [["/search", 10], ["/upload", 20]])

    row = ingest_table(xlsx_path, tmp_path, FakeEmbeddingClient())

    assert row.id == "tbl:latency.xlsx"


def test_ingest_table_xlsx_sets_modality_and_text_source(tmp_path):
    xlsx_path = tmp_path / "latency.xlsx"
    _write_xlsx(xlsx_path, ["a", "b"], [["1", "2"]])

    row = ingest_table(xlsx_path, tmp_path, FakeEmbeddingClient())

    assert row.modality == Modality.TABLE
    assert row.text_source == TextSource.TABLE_MARKDOWN


def test_ingest_table_xlsx_content_matches_equivalent_csv(tmp_path):
    header = ["endpoint", "p50_ms", "p95_ms"]
    data_rows = [["/search", "10", "20"], ["/upload", "30", "40"]]

    csv_path = tmp_path / "latency.csv"
    csv_lines = [",".join(header)] + [",".join(row) for row in data_rows]
    csv_path.write_text("\n".join(csv_lines) + "\n")

    xlsx_path = tmp_path / "latency.xlsx"
    _write_xlsx(xlsx_path, header, [[cell for cell in row] for row in data_rows])

    csv_row = ingest_table(csv_path, tmp_path, FakeEmbeddingClient())
    xlsx_row = ingest_table(xlsx_path, tmp_path, FakeEmbeddingClient())

    assert xlsx_row.content_text == csv_row.content_text


def test_ingest_table_xlsx_stringifies_numeric_and_none_cells(tmp_path):
    xlsx_path = tmp_path / "mixed.xlsx"
    # openpyxl yields typed values (int/float) and None for empty cells --
    # _rows_to_markdown does str.join, which crashes on non-str cells if the
    # xlsx reader doesn't stringify them first.
    _write_xlsx(xlsx_path, ["id", "value", "note"], [[1, 3.5, None], [2, None, "ok"]])

    row = ingest_table(xlsx_path, tmp_path, FakeEmbeddingClient())

    assert "| 1 | 3.5 |  |" in row.content_text
    assert "| 2 |  | ok |" in row.content_text


def test_ingest_table_xlsx_metadata_has_correct_row_and_column_counts(tmp_path):
    xlsx_path = tmp_path / "t.xlsx"
    _write_xlsx(xlsx_path, ["a", "b", "c"], [["1", "2", "3"], ["4", "5", "6"]])

    row = ingest_table(xlsx_path, tmp_path, FakeEmbeddingClient())

    metadata = json.loads(row.metadata)
    assert metadata["n_rows"] == 2
    assert metadata["n_cols"] == 3
    assert metadata["columns"] == ["a", "b", "c"]


def test_ingest_table_xlsx_reads_only_the_active_sheet(tmp_path):
    xlsx_path = tmp_path / "multi_sheet.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "first"
    ws1.append(["a", "b"])
    ws1.append(["1", "2"])
    ws2 = wb.create_sheet("second")
    ws2.append(["x", "y"])
    ws2.append(["should", "not-appear"])
    wb.save(xlsx_path)

    row = ingest_table(xlsx_path, tmp_path, FakeEmbeddingClient())

    assert "should" not in row.content_text
    assert "not-appear" not in row.content_text
